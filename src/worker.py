# -*- coding: utf-8 -*-
"""
招商银行交易流水提取 - 控制台工作进程
由 C# GUI 调用，通过 stdout 输出进度，exit code 表示结果。

用法: worker.exe <pdf_path> <output_xlsx_path>
"""

import os
import re
import sys
import io

# 强制 stdout/stderr 使用 UTF-8（Windows 控制台默认 GBK，C# 端按 UTF-8 读取）
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'buffer'):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from openpyxl import Workbook
from pdf_cmap_decoder import extract_text


def parse_transactions(pages_lines):
    transactions = []
    date_pat = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    amount_pat = re.compile(r'^-?[\d,]+\.\d{2}$')
    page_num_pat = re.compile(r'^\d+/\d+$')

    skip_kw = {
        'Transaction Statement', 'Name', 'Account Type',
        'Account No', 'Sub Branch', 'Verification Code',
        'Date', 'Currency', 'Transaction', 'Amount',
        'Balance', 'Counter Party', 'Transaction Type',
    }
    header_kw = [
        '记账日期', '货币', '交易金额', '联机余额',
        '交易摘要', '对手信息', '招商银行交易流水',
        '户  名', '账户类型', '申请时间', '账号',
        '开 户 行', '验 证 码',
    ]

    rec = None
    state = 0
    t_date = t_amt = t_bal = None

    for lines in pages_lines:
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if page_num_pat.match(line):
                continue
            if line in skip_kw:
                continue
            if any(k in line for k in header_kw):
                continue
            if '--' in line and re.search(r'\d{4}-\d{2}-\d{2}', line):
                continue

            if state == 0:
                if date_pat.match(line):
                    t_date = line; state = 1
                elif rec:
                    rec['对手信息'] = (rec['对手信息'] + ' ' + line).strip()
            elif state == 1:
                state = 2 if line == 'CNY' else 0
            elif state == 2:
                if amount_pat.match(line):
                    t_amt = float(line.replace(',', '')); state = 3
                else:
                    state = 0
            elif state == 3:
                if amount_pat.match(line):
                    t_bal = float(line.replace(',', '')); state = 4
                else:
                    state = 0
            elif state == 4:
                if date_pat.match(line):
                    if rec: transactions.append(rec)
                    rec = {'记账日期': t_date, '币种': 'CNY',
                           '交易金额': t_amt, '联机余额': t_bal,
                           '交易摘要': '', '对手信息': ''}
                    t_date = line; state = 1
                else:
                    if rec: transactions.append(rec)
                    rec = {'记账日期': t_date, '币种': 'CNY',
                           '交易金额': t_amt, '联机余额': t_bal,
                           '交易摘要': line, '对手信息': ''}
                    state = 5
            elif state == 5:
                if date_pat.match(line):
                    t_date = line; state = 1
                elif line not in skip_kw:
                    rec['对手信息'] = (rec['对手信息'] + ' ' + line).strip()

    if rec:
        transactions.append(rec)
    return transactions


def export_to_excel(transactions, output_path):
    for t in transactions:
        for k in t:
            if isinstance(t[k], str):
                t[k] = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', t[k])

    cols = ['记账日期', '币种', '交易金额', '联机余额', '交易摘要', '对手信息']
    wb = Workbook()
    ws = wb.active
    ws.title = '交易流水'
    ws.append(cols)
    for t in transactions:
        ws.append([t[c] for c in cols])
    for col, wd in {'A': 14, 'B': 8, 'C': 14, 'D': 14, 'E': 22, 'F': 55}.items():
        ws.column_dimensions[col].width = wd
    wb.save(output_path)
    return len(transactions)


# ============================================================
#  信用卡对账单解析
# ============================================================

def detect_format(pages_lines):
    """判断 PDF 类型：'credit_card'（信用卡对账单）或 'transaction'（交易流水）。"""
    for page in pages_lines:
        for ln in page:
            if '信用卡对账单' in ln or '本期账务明细' in ln or 'Transaction Details' in ln:
                return 'credit_card'
    return 'transaction'


def parse_credit_card(pages_lines):
    """解析招商银行信用卡对账单交易明细。

    原生提取后每条交易为连续多行：交易日/记账日/交易摘要/人民币金额/卡号末四位/交易地金额，
    并由分类标记（还款/退款/消费/其他）分组。
    """
    date_pat = re.compile(r'^\d{2}/\d{2}$')
    amount_pat = re.compile(r'^-?[\d,]+\.\d{2}$')
    card4_pat = re.compile(r'^\d{4}$')
    categories = {'还款', '退款', '消费', '其他'}

    lines = [ln.strip() for page in pages_lines for ln in page]

    txns = []
    cur_cat = ''
    i, n = 0, len(lines)
    while i < n:
        line = lines[i]
        if not line:
            i += 1
            continue
        if line in categories:
            cur_cat = line
            i += 1
            continue
        if date_pat.match(line):
            sold = line
            posted = ''
            desc_parts = []
            rmb = card4 = orig = ''
            j = i + 1
            if j < n and date_pat.match(lines[j]):
                posted = lines[j]
                j += 1
            # 交易摘要：直到遇到金额行（或下一个日期/分类）
            while j < n and lines[j] and not amount_pat.match(lines[j]) \
                    and not date_pat.match(lines[j]) and lines[j] not in categories:
                desc_parts.append(lines[j])
                j += 1
            desc = ' '.join(desc_parts)
            if j < n and amount_pat.match(lines[j]):
                rmb = lines[j]
                j += 1
            if j < n and card4_pat.match(lines[j]):
                card4 = lines[j]
                j += 1
            if j < n and amount_pat.match(lines[j]):
                orig = lines[j]
                j += 1
            txns.append({
                '交易分类': cur_cat,
                '交易日': sold,
                '记账日': posted,
                '交易摘要': desc,
                '人民币金额': rmb,
                '卡号末四位': card4,
                '交易地金额': orig,
            })
            i = j
            continue
        i += 1
    return txns


def extract_card_summary(pages_lines):
    """从信用卡账单中提取账单概览信息（best-effort，取不到则留空）。"""
    text = '\n'.join(ln for page in pages_lines for ln in page)
    info = {'账单日': '', '到期还款日': '', '信用额度': '', '本期还款总额': '', '本期最低还款额': ''}

    m = re.search(r'账单日[^\d]*(\d{4}年\d{1,2}月\d{1,2}日)', text)
    if m:
        info['账单日'] = m.group(1)
    m = re.search(r'到期还款日[^\d]*(\d{4}年\d{1,2}月\d{1,2}日)', text)
    if m:
        info['到期还款日'] = m.group(1)
    m = re.search(r'信用额度[^\d¥]*¥?\s*([\d,]+\.\d{2})', text)
    if m:
        info['信用额度'] = m.group(1)
    # 本期还款总额 与 本期最低还款额 为标签后紧邻的两个金额
    m = re.search(r'本期还款总额.*?¥?\s*([\d,]+\.\d{2}).*?¥?\s*([\d,]+\.\d{2})', text, re.DOTALL)
    if m:
        info['本期还款总额'] = m.group(1)
        info['本期最低还款额'] = m.group(2)
    return info


def _to_float(s):
    try:
        return float(str(s).replace(',', '').strip())
    except (ValueError, TypeError):
        return s


def _clean(s):
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', str(s))


def _full_date(mmdd, stmt_year, stmt_month):
    """将 MM/DD 补全为 YYYY-MM-DD；年份取自账单日，跨年（如账单1月、交易12月）自动修正为上一年。"""
    m = re.match(r'^(\d{2})/(\d{2})$', str(mmdd).strip())
    if not m or not stmt_year:
        return mmdd
    month, day = int(m.group(1)), int(m.group(2))
    year = stmt_year
    if stmt_month and month > stmt_month + 1:
        year -= 1
    return f'{year:04d}-{month:02d}-{day:02d}'


def export_credit_card_excel(txns, summary, output_path, stmt_year=None, stmt_month=None):
    cols = ['交易分类', '交易日', '记账日', '交易摘要', '人民币金额', '卡号末四位', '交易地金额']
    wb = Workbook()
    ws = wb.active
    ws.title = '交易明细'
    ws.append(cols)
    for t in txns:
        ws.append([
            t['交易分类'],
            _full_date(t['交易日'], stmt_year, stmt_month),
            _full_date(t['记账日'], stmt_year, stmt_month),
            _clean(t['交易摘要']),
            _to_float(t['人民币金额']),
            t['卡号末四位'],
            _to_float(t['交易地金额']),
        ])
    for col, wd in {'A': 10, 'B': 12, 'C': 12, 'D': 42, 'E': 14, 'F': 12, 'G': 14}.items():
        ws.column_dimensions[col].width = wd

    if summary:
        ws2 = wb.create_sheet('账单概览')
        for k, v in summary.items():
            ws2.append([k, v])
        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 24

    wb.save(output_path)
    return len(txns)


def log(msg):
    """输出一行日志到 stdout（供 GUI 读取）"""
    print(msg, flush=True)


def main():
    if len(sys.argv) < 3:
        log("[ERROR] 用法: worker.exe <pdf路径> <输出xlsx路径>")
        sys.exit(1)

    pdf_path = sys.argv[1]
    output_path = sys.argv[2]

    if not os.path.isfile(pdf_path):
        log(f"[ERROR] PDF文件不存在: {pdf_path}")
        sys.exit(1)

    try:
        log(f"读取 PDF: {os.path.basename(pdf_path)}")
        log("正在解析 PDF 字体编码…")

        pages = extract_text(pdf_path)
        n = sum(len(p) for p in pages)
        log(f"提取 {len(pages)} 页, {n} 行文本")

        fmt = detect_format(pages)

        if fmt == 'credit_card':
            log("检测到信用卡对账单格式")
            summary = extract_card_summary(pages)
            log("正在解析交易记录…")
            txns = parse_credit_card(pages)
            log(f"解析 {len(txns)} 条交易记录")
            if not txns:
                log("[ERROR] 未解析到交易记录，请确认PDF为招商银行信用卡对账单格式。")
                sys.exit(1)
            m = re.search(r'(\d{4})年(\d{1,2})月', summary.get('账单日', ''))
            stmt_year = int(m.group(1)) if m else None
            stmt_month = int(m.group(2)) if m else None
            log("正在导出 Excel…")
            cnt = export_credit_card_excel(txns, summary, output_path, stmt_year, stmt_month)
            log(f"导出 {cnt} 条记录 → {output_path}")
            log(f"[DONE] {output_path}")
            sys.exit(0)

        log("检测到交易流水格式")
        log("正在解析交易记录…")
        txns = parse_transactions(pages)
        log(f"解析 {len(txns)} 条交易记录")

        if not txns:
            log("[ERROR] 未解析到交易记录，请确认PDF为招商银行交易流水格式。")
            sys.exit(1)

        log("正在导出 Excel…")
        cnt = export_to_excel(txns, output_path)
        log(f"导出 {cnt} 条记录 → {output_path}")
        log(f"[DONE] {output_path}")
        sys.exit(0)

    except Exception as exc:
        log(f"[ERROR] {exc}")
        sys.exit(1)


if __name__ == '__main__':
    main()
